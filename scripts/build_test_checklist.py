#!/usr/bin/env python3
# Generuje Cloak-and-Dagger-Checklista-Testow.xlsx — checklista do odhaczania,
# czy każda funkcjonalność wtyczki działa. Uruchom: python3 scripts/build_test_checklist.py
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

OUT = "Cloak-and-Dagger-Checklista-Testow.xlsx"

# ── Paleta (spójna z brandem wtyczki) ───────────────────────────────────────
INK      = "0A0B0E"   # tło nagłówka
ACCENT   = "2BD4C4"   # turkus
ACCENT2  = "1C8F86"
DANGER   = "E5484D"
PURPLE   = "9A8CFF"
GREY     = "6E7480"
LIGHT    = "F4F6F8"
ZEBRA    = "EEF6F5"
WHITE    = "FFFFFF"

def fill(c): return PatternFill("solid", fgColor=c)
def font(**k): return Font(name="Calibri", **k)
thin = Side(style="thin", color="D5DBDF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Dane testów: (ID, Moduł, Funkcjonalność, Kroki, Oczekiwany wynik) ────────
TESTS = [
    ("T01", "DataGhost", "Silnik szumu (ruch-wabik)",
     "SW console: chrome.runtime.sendMessage({type:'TRIGGER_NOISE'}) lub odczekaj cykl ~1 min. Otwórz Dashboard → Telemetria.",
     "Wpisy GHOST 'DataGhost: batch N zapytań'; licznik 'wstrzyknięć szumu' i kafelek RUCH-WABIK rosną."),
    ("T02", "Honeypot Trap", "Zatruwanie profilera (data poisoning)",
     "SW console: chrome.runtime.sendMessage({type:'TRIGGER_HONEYPOT_TEST'}). Albo wejdź na stronę z Google Analytics / Pixel FB.",
     "Telemetria TRAP 'Zatruto Google Analytics: …absurdalna persona'; licznik 'zatrutych trackerów' rośnie; kropka na radarze."),
    ("T03", "Honeypot Trap", "Rotacja trucizny w czasie",
     "Po teście T02 odczekaj 1 min (alarm honeypot-rotate-poison) i wywołaj test ponownie.",
     "Opis wstrzykniętego profilu w nowym logu jest INNY niż poprzednio."),
    ("T04", "Cookie Shredder", "Rotacja ciasteczek trackerów",
     "Wejdź na stronę z GA. DevTools strony: document.cookie → zanotuj _ga. Przełącz moduł OFF→ON (rotacja od razu) lub odczekaj 1 min. Odśwież document.cookie.",
     "Segment-ID w _ga zmieniony (ta sama długość/klasa znaków), prefiks GA1.1. nietknięty; log 'Zrotowano N ciasteczek'; kafelek CIASTECZKA ZROTOWANE rośnie."),
    ("T05", "Cookie Shredder", "Bezpieczeństwo — nie rusza logowania",
     "Zaloguj się gdzieś (ciasteczka SID/CONSENT/csrftoken). Wykonaj rotację jak w T04.",
     "Po rotacji pozostajesz ZALOGOWANY; ciasteczka sesji/zgód bez zmian."),
    ("T06", "Targeting Shield", "Strip atrybucji (gclid/fbclid/utm)",
     "Wejdź: https://example.com/?gclid=TEST123&utm_source=news&utm_medium=email&fbclid=ABC",
     "Parametry gclid/fbclid/utm_* znikają z URL po załadowaniu; (dev) kafelek ATRYBUCJA ZERWANA rośnie."),
    ("T07", "Targeting Shield", "Total blackout per-origin (ręcznie)",
     "Na wrażliwej stronie: SW console chrome.runtime.sendMessage({type:'TRIGGER_TARGETING_TEST'}). Odśwież.",
     "SW log 'BLACKOUT (test) dla originu'; żądania do doubleclick/facebook/criteo/taboola blokowane (Network=blocked); reguła DNR 43100+; host w cnd:targeting:blocked-origins."),
    ("T08", "Targeting Shield", "Blackout automatyczny z AI",
     "Wejdź na stronę wrażliwą; poczekaj na skan AI (T11) z poziomem high/critical.",
     "Origin trafia automatycznie do blackoutu (cnd:targeting:blocked-origins) bez ręcznego wyzwalania."),
    ("T09", "Bionic Blur", "Maskowanie myszy + klawiatury (MAIN world)",
     "Strona http(s) → DevTools Console: window.__cloakDaggerBionicBlurInstalled. Wpisz tekst w pole formularza.",
     "Zwraca true; pisanie działa płynnie (input nieuszkodzony). Smoke: npm run smoke:extension PASS."),
    ("T10", "Bionic Blur", "Spójny fingerprint (navigator.*)",
     "Po aktywacji Wirtualnej Tożsamości (T13) porównaj navigator.hardwareConcurrency / userAgent z czystą kartą bez wtyczki.",
     "Wartości podmienione i WEWNĘTRZNIE spójne (UA pasuje do platformy)."),
    ("T11", "AI Deep-Dive", "Lokalna detekcja ryzyka strony",
     "Wejdź na stronę o wrażliwej treści. SW console: chrome.storage.local.get('cnd:state').then(s=>console.log(s['cnd:state'].aiDeepDiveRisk, s['cnd:state'].maxCamoActive))",
     "aiDeepDiveRisk ustawione (level+score); przy wysokim ryzyku maxCamoActive=true i zazbrojenie DataGhost/Mouse/Keystroke."),
    ("T12", "Cień cyfrowy (Shadow Audit)", "Profil z realnej historii",
     "Dashboard → prawa kolumna → rozwiń 'Cień cyfrowy · audyt' → skanuj. Wcześniej odwiedź kilka stron jednej kategorii (np. github, stackoverflow).",
     "Pasek 'Twój ślad' vs 'Maska' + drop w bitach; sekcja Zainteresowania z DOWODAMI (konkretne domeny); zgadywanka płci/wieku z niską pewnością. Brak żądań sieciowych."),
    ("T13", "Wirtualna Tożsamość", "Kreator + aktywacja profilu",
     "Dashboard → prawa krawędź → uchwyt 'Wirtualna Tożsamość'. Wybierz archetyp/parametry → Aktywuj.",
     "Tylko jedna lista archetypów (bez zakładki 'Specjalne'); log 'Aktywowano tożsamość…'; zapis cnd:virtual-identity:active, profile-id='custom', cnd:dataghost:topics."),
    ("T14", "Panic Button", "Strefa awaryjna (kill-switch)",
     "Dashboard → środek pod radarem → Strefa awaryjna → PRZYTRZYMAJ ~0,85 s.",
     "Komunikat 'wyczyszczone'; cnd:state liczniki=0; telemetria (cnd:logs) pusta; realnie skasowane cookies/cache (wylogowanie). Puszczenie wcześniej = anulowanie."),
    ("T15", "Alias e-mail", "Identity masking",
     "Dashboard/popup → Generuj alias e-mail; potem 'nowy'.",
     "Pojawia się alias …@…; log 'Wygenerowano alias'; online=API SimpleLogin, offline=fallback bez crasha."),
    ("T16", "Privacy Score", "Spójność i przeliczanie",
     "Włączaj/wyłączaj moduły i generuj aktywność. Obserwuj Privacy Score w popupie i dashboardzie.",
     "Score zmienia się na żywo i jest IDENTYCZNY w popupie i dashboardzie (STATE_UPDATE)."),
    ("T17", "Persystencja telemetrii", "Logi przeżywają zamknięcie okna",
     "Wygeneruj kilka logów → zamknij Dashboard → otwórz ponownie.",
     "Telemetria nadal jest (cnd:logs, do 50 wpisów). Powtarzalne zdarzenia łączą się w ×N."),
    ("T18", "Synchronizacja popup↔dashboard", "Stan współdzielony",
     "Otwórz popup i Dashboard jednocześnie; zmień coś w jednym.",
     "Zmiana natychmiast widoczna w drugim oknie."),
    ("T19", "Layout dashboardu", "Nowy układ 3 kolumn",
     "Otwórz Dashboard na pełnym ekranie.",
     "Lewa: Score+kafelki. Środek: Radar + (pod nim) Strefa awaryjna + Generuj alias. Prawa: Wektory ochrony + Telemetria + Cień cyfrowy. BRAK karty AI Deep-Dive. Brak nakładania elementów."),
    ("T20", "Testy automatyczne", "CI lokalne",
     "npm run typecheck && npm test && npm run build",
     "typecheck 0 błędów; vitest 78/78 PASS; build prod kończy się sukcesem."),
]

STATUSES = ["", "✅ DZIAŁA", "⚠️ CZĘŚCIOWO", "❌ NIE DZIAŁA", "⏭️ POMINIĘTE"]
STATUS_COLORS = {
    "✅ DZIAŁA": "C8F2E0",
    "⚠️ CZĘŚCIOWO": "FDEFC2",
    "❌ NIE DZIAŁA": "F8CDCE",
    "⏭️ POMINIĘTE": "E3E6EA",
}

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════════
# Arkusz 1: Checklista
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Checklista testów"
ws.sheet_view.showGridLines = False

# Tytuł
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "CLOAK & DAGGER — CHECKLISTA TESTÓW FUNKCJONALNOŚCI"
t.font = font(bold=True, size=16, color=ACCENT)
t.fill = fill(INK)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:H2")
s = ws["A2"]
s.value = ("Odhaczaj kolumnę STATUS dla każdego testu. Szczegóły kroków/snippetów: TESTING.md  ·  "
           f"Wygenerowano: {datetime.date.today().isoformat()}")
s.font = font(size=9, italic=True, color=GREY)
s.fill = fill(LIGHT)
s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

# Nagłówek tabeli
headers = ["ID", "Moduł", "Funkcjonalność", "Jak sprawdzić (kroki)",
           "Oczekiwany wynik (PASS gdy…)", "STATUS", "Tester", "Uwagi"]
HEAD_ROW = 4
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=HEAD_ROW, column=i, value=h)
    c.font = font(bold=True, size=10, color=WHITE)
    c.fill = fill(ACCENT2)
    c.alignment = center
    c.border = border
ws.row_dimensions[HEAD_ROW].height = 26

# Wiersze danych
first = HEAD_ROW + 1
for r, (tid, mod, func, steps, expected) in enumerate(TESTS, start=first):
    vals = [tid, mod, func, steps, expected, "", "", ""]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.border = border
        c.alignment = center if i in (1, 6, 7) else wrap
        if i == 1:
            c.font = font(bold=True, size=10, color=ACCENT2)
        elif i == 2:
            c.font = font(bold=True, size=9)
        else:
            c.font = font(size=9)
    # zebra
    if (r - first) % 2 == 1:
        for i in range(1, 9):
            ws.cell(row=r, column=i).fill = fill(ZEBRA)
    ws.row_dimensions[r].height = 58
last = first + len(TESTS) - 1

# Szerokości kolumn
widths = {"A": 6, "B": 20, "C": 26, "D": 52, "E": 52, "F": 16, "G": 12, "H": 24}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Walidacja statusu (dropdown) + kolorowanie warunkowe
dv = DataValidation(type="list", formula1='"{}"'.format(",".join(STATUSES[1:])), allow_blank=True)
dv.prompt = "Wybierz status testu"
dv.promptTitle = "Status"
ws.add_data_validation(dv)
dv.add(f"F{first}:F{last}")

from openpyxl.formatting.rule import CellIsRule
for label, col in STATUS_COLORS.items():
    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        CellIsRule(operator="equal", formula=[f'"{label}"'],
                   fill=fill(col), font=font(bold=True, size=9)))

ws.freeze_panes = f"A{first}"
ws.auto_filter.ref = f"A{HEAD_ROW}:H{last}"

# ── Pasek podsumowania (liczniki przez COUNTIF) ─────────────────────────────
sumrow = last + 2
ws.merge_cells(f"A{sumrow}:C{sumrow}")
cell = ws[f"A{sumrow}"]
cell.value = "PODSUMOWANIE"
cell.font = font(bold=True, size=11, color=WHITE)
cell.fill = fill(INK)
cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[sumrow].height = 22

summary = [
    ("✅ Działa",     f'=COUNTIF(F{first}:F{last},"✅ DZIAŁA")',        "C8F2E0"),
    ("⚠️ Częściowo",  f'=COUNTIF(F{first}:F{last},"⚠️ CZĘŚCIOWO")',     "FDEFC2"),
    ("❌ Nie działa", f'=COUNTIF(F{first}:F{last},"❌ NIE DZIAŁA")',     "F8CDCE"),
    ("⏭️ Pominięte",  f'=COUNTIF(F{first}:F{last},"⏭️ POMINIĘTE")',     "E3E6EA"),
    ("Σ Wszystkich",  f'=COUNTA(A{first}:A{last})',                     "D8F3EE"),
]
row = sumrow + 1
for label, formula, color in summary:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = font(bold=True, size=10)
    lc.fill = fill(color)
    lc.border = border
    lc.alignment = Alignment(indent=1, vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    vc = ws.cell(row=row, column=3, value=formula)
    vc.font = font(bold=True, size=11, color=ACCENT2)
    vc.fill = fill(color)
    vc.border = border
    vc.alignment = center
    row += 1

# Pasek postępu (% przetestowanych z wynikiem DZIAŁA)
prow = row + 1
ws.merge_cells(start_row=prow, start_column=1, end_row=prow, end_column=2)
pc = ws.cell(row=prow, column=1, value="% DZIAŁA (z odhaczonych)")
pc.font = font(bold=True, size=10)
pc.alignment = Alignment(indent=1, vertical="center")
pf = ws.cell(row=prow, column=3,
             value=f'=IFERROR(COUNTIF(F{first}:F{last},"✅ DZIAŁA")/'
                   f'(COUNTA(A{first}:A{last})-COUNTBLANK(F{first}:F{last})),0)')
pf.number_format = "0%"
pf.font = font(bold=True, size=12, color=ACCENT2)
pf.alignment = center

# ════════════════════════════════════════════════════════════════════════════
# Arkusz 2: Snippety diagnostyczne
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Snippety SW console")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:B1")
h = ws2["A1"]
h.value = "SNIPPETY DO KONSOLI SERVICE WORKERA (chrome://extensions → Sprawdź widoki: service worker)"
h.font = font(bold=True, size=12, color=ACCENT)
h.fill = fill(INK)
h.alignment = Alignment(indent=1, vertical="center")
ws2.row_dimensions[1].height = 30

snippets = [
    ("Storage dump (cały stan)", "chrome.storage.local.get(null).then(s=>console.log(JSON.stringify(s,null,2)))"),
    ("Stan dashboardu (Score+liczniki)", "chrome.storage.local.get('cnd:state').then(console.log)"),
    ("Podgląd reguł DNR (42001/43001/43100+)", "chrome.declarativeNetRequest.getDynamicRules().then(r=>console.table(r.map(x=>({id:x.id,type:x.action.type}))))"),
    ("Wyzwól szum (DataGhost)", "chrome.runtime.sendMessage({type:'TRIGGER_NOISE'})"),
    ("Wyzwól zatrucie (Honeypot)", "chrome.runtime.sendMessage({type:'TRIGGER_HONEYPOT_TEST'})"),
    ("Wyzwól blackout (Targeting bieżącej karty)", "chrome.runtime.sendMessage({type:'TRIGGER_TARGETING_TEST'})"),
    ("Chronione originy (blackout)", "chrome.storage.local.get('cnd:targeting:blocked-origins').then(console.log)"),
    ("Aktywna wirtualna tożsamość", "chrome.storage.local.get(['cnd:virtual-identity:active','cnd:bionic-blur:profile-id','cnd:dataghost:topics']).then(console.log)"),
    ("Telemetria (zapisane logi)", "chrome.storage.local.get('cnd:logs').then(console.log)"),
    ("Logi offscreen (model AI)", "chrome.storage.local.get('cnd:offscreen-logs').then(console.log)"),
    ("Bionic Blur wstrzyknięty? (DevTools STRONY)", "window.__cloakDaggerBionicBlurInstalled"),
]
ws2.cell(row=3, column=1, value="Cel").font = font(bold=True, color=WHITE)
ws2.cell(row=3, column=1).fill = fill(ACCENT2)
ws2.cell(row=3, column=2, value="Komenda").font = font(bold=True, color=WHITE)
ws2.cell(row=3, column=2).fill = fill(ACCENT2)
ws2.cell(row=3, column=1).border = border
ws2.cell(row=3, column=2).border = border
for r, (label, code) in enumerate(snippets, start=4):
    a = ws2.cell(row=r, column=1, value=label)
    a.font = font(bold=True, size=9)
    a.alignment = wrap
    a.border = border
    b = ws2.cell(row=r, column=2, value=code)
    b.font = Font(name="Consolas", size=9)
    b.alignment = wrap
    b.border = border
    if (r - 4) % 2 == 1:
        a.fill = fill(ZEBRA); b.fill = fill(ZEBRA)
    ws2.row_dimensions[r].height = 30
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 90

# ════════════════════════════════════════════════════════════════════════════
# Arkusz 3: Środowisko / setup
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Setup")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:B1")
hh = ws3["A1"]
hh.value = "PRZYGOTOWANIE ŚRODOWISKA TESTOWEGO"
hh.font = font(bold=True, size=12, color=ACCENT)
hh.fill = fill(INK)
hh.alignment = Alignment(indent=1, vertical="center")
ws3.row_dimensions[1].height = 30
steps = [
    ("1. Build dev", "npm run dev  → katalog build/chrome-mv3-dev (hot-reload, liczniki DNR aktywne)"),
    ("1b. Build prod", "npm run build → build/chrome-mv3-prod (wersja do oceny)"),
    ("2. Wczytaj wtyczkę", "chrome://extensions → Tryb dewelopera → Wczytaj rozpakowane → wskaż build/chrome-mv3-*"),
    ("3. Otwórz SW console", "Szczegóły wtyczki → Sprawdź widoki: service worker"),
    ("4. Otwórz Dashboard", "Kliknij ikonę wtyczki → ikona pełnego ekranu (Maximize) w nagłówku popupu"),
    ("WAŻNE: liczniki", "Liczniki Honeypot/Targeting rosną przez onRuleMatchedDebug = TYLKO wersja rozpakowana. Filtrowanie/zatruwanie działa też w prod."),
    ("WAŻNE: DataGhost", "To ruch-wabik (no-cors, bez ciasteczek) — szum sieciowy, NIE czyści profilu. Profilu dotyka Cookie Shredder."),
    ("WAŻNE: Cookie Shredder", "Rusza tylko ciasteczka czysto-trackingowe; po rotacji MUSISZ pozostać zalogowany (test T05)."),
]
for r, (k, v) in enumerate(steps, start=3):
    a = ws3.cell(row=r, column=1, value=k)
    a.font = font(bold=True, size=10, color=(DANGER if k.startswith("WAŻNE") else ACCENT2))
    a.alignment = wrap
    a.border = border
    b = ws3.cell(row=r, column=2, value=v)
    b.font = font(size=10)
    b.alignment = wrap
    b.border = border
    if k.startswith("WAŻNE"):
        a.fill = fill("FDEFC2"); b.fill = fill("FDEFC2")
    elif (r - 3) % 2 == 1:
        a.fill = fill(ZEBRA); b.fill = fill(ZEBRA)
    ws3.row_dimensions[r].height = 34
ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 96

wb.save(OUT)
print(f"OK → {OUT}  ({len(TESTS)} testów, 3 arkusze)")
